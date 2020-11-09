#Code used to create CORTES with Python

#==============================================================================
#Libraries.
#==============================================================================
from os import path
from openpyxl import Workbook, load_workbook
from pandas import DataFrame
#import win32com.client as w32lc
#import os
#import pandas as pd

#Paths
thisFilePath = path.abspath(__file__)
thisFolderPath = path.dirname(thisFilePath)

#==============================================================================
#Class definition.
#==============================================================================
class createCortes:
    def __init__(self, month, year, trimester, dictName):
        self.month = month
        self.year = year
        self.trimester = trimester
        self.dictName = dictName
    
        #function used to split the text into different tags
    def separateTags(self, text, MWlimit):
        text=text.replace(" ", "")  #removes all the spaces
        cortes = []
        iCorte=0
        cortes.append([])
        for Letter, iLetter in zip(text, range(len(text))):
            #condition used to remove sobrecargas and sobretensiones
            if MWlimit==None or MWlimit=="-" or MWlimit==" ":
                iNumTags ="Null"
                return cortes, iNumTags
            
            # "/" and "+" are commonly used to separate Tags
            if Letter == "/" or Letter == "+":
                # Checks if the "/" is for transformators ratio (and not to separate Tags)
                if Letter == "/" and text[iLetter-1].isdigit() and text[iLetter+1].isdigit():
                    cortes[iCorte].append(Letter)
                    continue
                # If it is "+" or "/" to separate Tags
                else:
                    iCorte+=1
                    cortes.append([])
                    continue
            #if the letter is not "/" nor "+"
            cortes[iCorte].append(Letter)   #appends the letter to the string
        iNumTags = iCorte+1
        return cortes, iNumTags

        #function used to split the text into different tags
    def separateTags2(self, text, MWlimit):
        text=text.replace(" ", "")  #removes all the spaces
        cortes = []
        iCorte=0
        cortes.append('')

        #condition used to remove sobrecargas and sobretensiones
        if MWlimit==None or MWlimit=="-" or MWlimit==" ":
            iNumTags ="Null"
            return cortes, iNumTags 

        for Letter, iLetter in zip(text, range(len(text))):
            # "/" and "+" are commonly used to separate Tags
            if Letter == "/" or Letter == "+":
                # Checks if the "/" is for transformators ratio (and not to separate Tags)
                if Letter == "/" and text[iLetter-1].isdigit() and text[iLetter+1].isdigit():
                    cortes[iCorte]+=Letter  #appends the letter to the string
                    continue
                # If it is "+" or "/" to separate Tags
                else:
                    iCorte+=1
                    cortes.append('')   #Creates new string
                    continue
            #if the letter is not "/" nor "+"
            cortes[iCorte]+=Letter   #appends the letter to the string
        iNumTags = iCorte+1
        return cortes, iNumTags
    
    #Function that looks for Corte in Dictionary and returns the related tags
    def findTag(self, key, maxCol, wsDict):
        P = 'NOT FOUND'
        Pqc = 'NOT FOUND'
        Q = 'NOT FOUND'
        Qqc = 'NOT FOUND'
        #remove the special characters in key string
        key=key.replace(" ","").replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u").replace("\n","").upper().replace("KV","").replace(".","").replace("-","").replace("–","")
        #Looks for the key in the Dictionary
        for iFila in range(2,maxCol+1):
            value=(str(wsDict.cell(row=iFila,column=1).value).replace(" ","").replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u").replace("\n","").upper()).replace("KV","").replace(".","").replace("-","").replace("–","")
            # If find the key, take the tags from the dictionary
            if key==value:
                P=wsDict.cell(row=iFila,column=2).value
                Pqc=wsDict.cell(row=iFila,column=3).value
                Q=wsDict.cell(row=iFila,column=4).value
                Qqc=wsDict.cell(row=iFila,column=5).value
                break
        return P,Pqc,Q,Qqc

    def writeCSV(self, sP, sPqc, sQ, sQqc, iCorte, iCorteValue, sArea, dfData):
        dfData = dfData.append({'P':sP, 'Pcalidad':sPqc, 'Q':sQ, 'Qcalidad':sQqc, 'SubArea':sArea, 'Corte':iCorte, 'Pmax':iCorteValue}, ignore_index=True)
        return dfData

#=================================================================================
    def runCreateCortes(self): 
#=================================================================================
        fileName = str(self.year) + '-' + str(self.month) + '_Restricciones_'+str(self.year)+'_T'+str(self.trimester)+'.xlsx'
        wbRestric = load_workbook(thisFolderPath+'\\'+fileName)     #Load Restricciones workbook
        wbDict = load_workbook(thisFolderPath+'\\'+self.dictName)        #Load Diccionario workbook
        dfData = DataFrame([],columns=['P','Pcalidad','Q','Qcalidad','SubArea','Corte','Pmax'])
        wsRestric = wbRestric.active        #Get active worksheet on Restricciones workbook
        wsDict = wbDict.active              #Get active worksheet on Diccionario workbook
        maxColumnRest = wsRestric.max_row   #Get the max column value on Restricciones table
        iCantCortes = 0                     #Counter of Cortes to mark in the .csv file

        # Iterates over all rows in Restrictions ws
        for iRow in range(2, maxColumnRest+1):
            text = wsRestric.cell(row=iRow, column=3).value
            #Cortes, iNumTags = self.separateTags(text, wsRestric.cell(row=iRow, column=7).value) #gets original strings
            Cortes2, iNumTags2 = self.separateTags2(text, wsRestric.cell(row=iRow, column=7).value) #gets strings concatenated
            # If MWlimit is Null, it is not considered
            if iNumTags2 == "Null":
                continue
            # If MWlimit is different to Null, counts a new Corte
            iCantCortes+=1
            for iCorte in Cortes2:
                #Looks for tag in dictionary
                sP, sPqc, sQ, sQqc = self.findTag(iCorte, wsDict.max_row, wsDict)
                iCorteValue = wsRestric.cell(row=iRow, column=7).value
                sArea = wsRestric.cell(row=iRow, column=1).value
                dfData = self.writeCSV(sP, sPqc, sQ, sQqc, iCantCortes, iCorteValue, sArea, dfData) 
            wsRestric.cell(row=iRow, column=8).value=str(iCantCortes)
        
        wbRestric.save(thisFolderPath+'\\'+fileName) #Save the restrictions workbook
        dfData.to_csv(thisFolderPath+'\\'+'Cortes_creados_'+self.year+'-'+self.month+'.csv', sep=',', index=False) #Save the dfData to a CSV file
#==============================================================================
#Main.
#==============================================================================
def main():
    
    Cortes=createCortes(month, year, trimester, dictName)       #Creates class named Cortes
    Cortes.runCreateCortes()    #Runs the function that creates cortes


#==============================================================================
#CODE STARTS HERE!!!!!
#==============================================================================
if __name__ == "__main__":
    #-----------------------------------------
    # Preliminars.
    month = '03'           #Month to analyse
    year = '2020'         #Year to analyse
    trimester = (int(month)-1)//3 +1        #Trimester of the year
    dictName = 'diccionario.xlsx'           #Dictionary filename

    #-----------------------------------------
    # run main.
    main()